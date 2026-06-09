import re
from decimal import Decimal
from io import BytesIO
from datetime import date

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from psycopg2.extras import execute_values


class FormulaError(Exception):
    pass


def fetch_df(cur, query, params=None):
    cur.execute(query, params or ())
    rows = cur.fetchall()
    cols = [desc[0] for desc in cur.description]
    return pd.DataFrame(rows, columns=cols)


def slug(value):
    return str(value or "").strip().lower().replace(" ", "_").replace("-", "_")


def label_from_key(key):
    return str(key).replace("_", " ").title()


def clean_number(value):
    if value in [None, ""]:
        return None

    try:
        if pd.isna(value):
            return None
    except Exception:
        pass

    try:
        return float(value)
    except Exception:
        return None


def clean_int(value):
    number = clean_number(value)

    if number is None:
        return 0

    return int(float(number))


def decimal_value(value):
    number = clean_number(value)

    if number is None:
        return Decimal("0")

    return Decimal(str(round(number, 2)))


def display_number(value):
    number = clean_number(value)

    if number is None:
        return ""

    if float(number).is_integer():
        return int(number)

    return round(number, 2)


def is_flush_component(component):
    component_key = slug(component)
    return (
        "flush_shutter" in component_key
        or "flush_door" in component_key
        or component_key == "flush"
    )


def component_sort_key(component):
    if is_flush_component(component):
        return 9999

    order = {
        "frame_vertical": 1,
        "frame_horizontal": 2,
        "architrave_vertical": 3,
        "architrave_horizontal": 4,
    }

    return order.get(slug(component), 100)


def wood_shade_for_component(component):
    if is_flush_component(component):
        return "38 mm BB"

    return "SOLIDWOOD - TEAK"


def product_uses_lh_rh(product_cat, product_code, rules):
    product_text = slug(f"{product_cat} {product_code}")

    if "door" in product_text or "shutter" in product_text:
        return True

    for rule in rules:
        formula_vars = extract_formula_variables(rule.get("formula_used"))

        if "lh_quantity" in formula_vars or "rh_quantity" in formula_vars:
            return True

    return False


def get_distinct_values(cur, table_name, column_name, where_sql="", params=None):
    cur.execute(
        f"""
        SELECT DISTINCT {column_name}
        FROM {table_name}
        {where_sql}
        ORDER BY {column_name}
        """,
        params or ()
    )
    return [row[0] for row in cur.fetchall() if row[0] is not None]


def load_product_rules(cur, product_cat, product_code):
    cur.execute(
        """
        SELECT
            product_cat,
            product_code,
            component,
            attribute,
            "type" AS rule_type,
            formula_used,
            fixed_value,
            quantity,
            display_order
        FROM product_component_rules
        WHERE product_cat = %s
          AND product_code = %s
        ORDER BY component, display_order
        """,
        (product_cat, product_code)
    )

    col_names = [desc[0] for desc in cur.description]
    return [dict(zip(col_names, row)) for row in cur.fetchall()]


def calculated_variable_keys(rules):
    keys = set()

    for rule in rules:
        component = slug(rule["component"])
        attribute = slug(rule["attribute"])

        keys.add(component)
        keys.add(f"{component}_{attribute}")

    return keys


def extract_formula_variables(formula):
    formula = str(formula or "")

    if not formula.strip():
        return set()

    names = re.findall(r"\b[A-Za-z_][A-Za-z0-9_]*\b", formula)

    ignored = {
        "abs",
        "min",
        "max",
        "round",
        "int",
        "float",
    }

    return {
        slug(name)
        for name in names
        if slug(name) not in ignored
    }


def required_formula_inputs(rules):
    inputs = set()

    for rule in rules:
        rule_type = str(rule["rule_type"] or "").lower()

        if rule_type == "formula":
            inputs.update(extract_formula_variables(rule["formula_used"]))

    inputs -= calculated_variable_keys(rules)

    inputs.discard("lh_quantity")
    inputs.discard("rh_quantity")
    inputs.discard("quantity")

    return sorted(inputs)


def required_manual_inputs(rules):
    inputs = []

    for rule in rules:
        rule_type = str(rule["rule_type"] or "").lower()
        attribute = slug(rule["attribute"])

        if rule_type != "manual":
            continue

        if attribute == "quantity":
            continue

        key = f"manual_{slug(rule['component'])}_{attribute}"
        label = f"{rule['component']} {rule['attribute']}"

        inputs.append({
            "key": key,
            "label": label,
            "component": rule["component"],
            "attribute": rule["attribute"],
        })

    return inputs


def safe_eval_formula(formula, variables):
    formula = str(formula or "").strip()

    if not formula:
        raise FormulaError("Formula is empty.")

    allowed_names = {
        "abs": abs,
        "min": min,
        "max": max,
        "round": round,
    }

    clean_variables = {}

    for key, value in variables.items():
        clean_variables[slug(key)] = float(value)

    eval_scope = {}
    eval_scope.update(allowed_names)
    eval_scope.update(clean_variables)

    try:
        result = eval(formula, {"__builtins__": {}}, eval_scope)
        return Decimal(str(round(float(result), 2)))

    except NameError as e:
        missing = str(e)
        raise FormulaError(f"Missing input in formula: {missing}")

    except Exception as e:
        raise FormulaError(str(e))


def get_component_quantity(component_rules):
    for rule in component_rules:
        if slug(rule["attribute"]) != "quantity":
            continue

        qty = clean_number(rule.get("quantity"))

        if qty is not None:
            return int(qty)

        fixed_qty = clean_number(rule.get("fixed_value"))

        if fixed_qty is not None:
            return int(fixed_qty)

    return 1


def calculate_cft(length, width, thickness, qty, component=None):
    if is_flush_component(component):
        return Decimal("0")

    length_num = clean_number(length) or 0
    width_num = clean_number(width) or 0
    thickness_num = clean_number(thickness) or 0
    quantity_num = clean_number(qty) or 0

    if length_num <= 0 or width_num <= 0 or thickness_num <= 0 or quantity_num <= 0:
        return Decimal("0")

    cft = length_num * width_num * thickness_num / 1000000000 * 35.315 * quantity_num
    return Decimal(str(round(cft, 2)))


def store_calculated_variable(variables, component, attribute, value):
    component_key = slug(component)
    attribute_key = slug(attribute)

    variables[f"{component_key}_{attribute_key}"] = float(value)

    if attribute_key == "length":
        variables[component_key] = float(value)


def calculate_rules(rules, user_inputs):
    variables = dict(user_inputs)
    pending_rules = [
        rule for rule in rules
        if slug(rule["attribute"]) != "quantity"
    ]

    calculated = []
    loop_count = 0

    while pending_rules:
        loop_count += 1

        if loop_count > 50:
            raise FormulaError("Formula dependency loop found.")

        progressed = False
        next_pending = []

        for rule in pending_rules:
            component = rule["component"]
            attribute = rule["attribute"]
            rule_type = str(rule["rule_type"] or "").lower()

            try:
                if rule_type == "fixed":
                    value = decimal_value(rule["fixed_value"])

                elif rule_type == "manual":
                    manual_key = f"manual_{slug(component)}_{slug(attribute)}"
                    value = decimal_value(variables.get(manual_key))

                elif rule_type == "formula":
                    value = safe_eval_formula(rule["formula_used"], variables)

                else:
                    value = Decimal("0")

                store_calculated_variable(variables, component, attribute, value)

                calculated.append({
                    "component": component,
                    "attribute": attribute,
                    "type": rule_type,
                    "formula": rule["formula_used"],
                    "value": value,
                })

                progressed = True

            except FormulaError as e:
                if "Missing input" in str(e) or "name" in str(e).lower():
                    next_pending.append(rule)
                else:
                    raise

        if not progressed:
            missing_items = [
                f"{rule['component']} {rule['attribute']}"
                for rule in next_pending
            ]
            raise FormulaError(
                "Cannot calculate: " + ", ".join(missing_items)
            )

        pending_rules = next_pending

    return calculated, variables


def build_component_summary(rules, calculated_rows, product_qty):
    rules_by_component = {}

    for rule in rules:
        component = rule["component"]

        if component not in rules_by_component:
            rules_by_component[component] = []

        rules_by_component[component].append(rule)

    calculated_by_component = {}

    for row in calculated_rows:
        component = row["component"]
        attribute = slug(row["attribute"])

        if component not in calculated_by_component:
            calculated_by_component[component] = {}

        calculated_by_component[component][attribute] = row["value"]

    summary_rows = []

    for component, component_rules in rules_by_component.items():
        base_qty = get_component_quantity(component_rules)
        total_qty = int(base_qty * product_qty)

        values = calculated_by_component.get(component, {})

        length = values.get("length", Decimal("0"))
        width = values.get("width", Decimal("0"))
        thickness = values.get("thickness", Decimal("0"))

        cft = calculate_cft(length, width, thickness, total_qty, component)

        summary_rows.append({
            "Component": component,
            "Length": display_number(length),
            "Width": display_number(width),
            "Thickness": display_number(thickness),
            "Qty": total_qty,
            "CFT": cft,
        })

    return summary_rows


def aggregate_preview_rows(raw_rows):
    if not raw_rows:
        return pd.DataFrame()

    raw_df = pd.DataFrame(raw_rows)

    grouped_rows = []

    group_cols = [
        "Product",
        "Component",
        "Length",
        "Width",
        "Thickness",
    ]

    for _, group_df in raw_df.groupby(group_cols, dropna=False, sort=False):
        first_row = group_df.iloc[0].to_dict()
        component = first_row["Component"]
        total_qty = int(group_df["Qty"].sum())

        cft = calculate_cft(
            first_row["Length"],
            first_row["Width"],
            first_row["Thickness"],
            total_qty,
            component,
        )

        grouped_rows.append({
            "Product": first_row["Product"],
            "Component": component,
            "Length": first_row["Length"],
            "Width": first_row["Width"],
            "Thickness": first_row["Thickness"],
            "Qty": total_qty,
            "CFT": cft,
            "LH & RH Details": "",
        })

    grouped_rows = sorted(
        grouped_rows,
        key=lambda row: component_sort_key(row["Component"])
    )

    lh_rh_summary = []

    unique_houses = raw_df[
        ["House Number", "LH Quantity", "RH Quantity"]
    ].drop_duplicates()

    for _, row in unique_houses.iterrows():
        house_number = row["House Number"]
        lh_qty = clean_int(row["LH Quantity"])
        rh_qty = clean_int(row["RH Quantity"])

        parts = []

        if lh_qty > 0:
            parts.append(f"{lh_qty}L")

        if rh_qty > 0:
            parts.append(f"{rh_qty}R")

        if parts:
            lh_rh_summary.append(f"{house_number}: {', '.join(parts)}")

    if grouped_rows and lh_rh_summary:
        chunk_size = max(1, int((len(lh_rh_summary) + len(grouped_rows) - 1) / len(grouped_rows)))

        chunks = []

        for idx in range(0, len(lh_rh_summary), chunk_size):
            chunks.append(" | ".join(lh_rh_summary[idx:idx + chunk_size]))

        for idx, chunk in enumerate(chunks):
            if idx < len(grouped_rows):
                grouped_rows[idx]["LH & RH Details"] = chunk

    return pd.DataFrame(grouped_rows)


def apply_cell_style(cell, border, font=None, fill=None, alignment=None):
    cell.border = border

    if font:
        cell.font = font

    if fill:
        cell.fill = fill

    if alignment:
        cell.alignment = alignment


def build_wood_issue_excel(
    project_name,
    unit_type,
    product_code,
    prepared_by,
    total_qty,
    total_lh_qty,
    total_rh_qty,
    opening_length,
    opening_width,
    preview_df,
):
    output = BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Wood Issue Report"

    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    green_fill = PatternFill("solid", fgColor="C6EFCE")
    product_fill = PatternFill("solid", fgColor="DDE7E0")
    white_fill = PatternFill("solid", fgColor="FFFFFF")

    title_font = Font(name="Calibri", size=14, bold=True)
    header_font = Font(name="Calibri", size=11, bold=True)
    normal_font = Font(name="Calibri", size=11)
    item_font = Font(name="Times New Roman", size=16)
    product_font = Font(name="Times New Roman", size=20, bold=True)

    center = Alignment(horizontal="center", vertical="center")
    wrap_center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.merge_cells("A1:M1")
    ws["A1"] = "WOOD ISSUE REPORT"
    apply_cell_style(ws["A1"], border, title_font, green_fill, center)

    ws.merge_cells("A2:A5")
    ws["A2"] = "1"

    ws.merge_cells("B2:D2")
    ws["B2"] = "PROJECT NAME"
    ws.merge_cells("E2:G2")
    ws["E2"] = project_name
    ws.merge_cells("H2:K2")
    ws["H2"] = f"{project_name} {unit_type} - {product_code} Batch"

    ws.merge_cells("L2:M4")
    ws["L2"] = total_qty

    ws.merge_cells("B3:D3")
    ws["B3"] = "PROJECT CODE"
    ws.merge_cells("E3:G3")
    ws["E3"] = ""

    ws.merge_cells("B4:D4")
    ws["B4"] = "PREPARED DATE"
    ws.merge_cells("E4:G4")
    ws["E4"] = ""
    ws.merge_cells("H4:K4")
    ws["H4"] = date.today().strftime("%d-%m-%Y")

    ws.merge_cells("B5:D5")
    ws["B5"] = "ORDER qty"
    ws.merge_cells("E5:G5")
    ws["E5"] = ""
    ws["H5"] = total_qty
    ws.merge_cells("I5:K5")
    ws["I5"] = f"{total_lh_qty} LH / {total_rh_qty} RH" if total_lh_qty or total_rh_qty else total_qty
    ws.merge_cells("L5:M5")
    ws["L5"] = "Prepared By"

    ws.merge_cells("A6:G6")
    ws.merge_cells("H6:I6")
    ws["H6"] = "OPENING SIZE"
    ws["J6"] = display_number(opening_length)
    ws["K6"] = display_number(opening_width)
    ws.merge_cells("L6:M6")
    ws["L6"] = prepared_by

    ws["A7"] = "SL.\nNO."
    ws.merge_cells("B7:C7")
    ws["B7"] = "Item Code"
    ws.merge_cells("D7:D12")
    ws["D7"] = product_code
    ws.merge_cells("E7:G7")
    ws["E7"] = "WOOD SHADE"
    ws["H7"] = "LENGTH"
    ws["I7"] = "WIDTH"
    ws["J7"] = "THICK"
    ws["K7"] = "QTY"
    ws["L7"] = "CFT"
    ws["M7"] = "LH & RH DETAILS"

    for row in range(1, 8):
        for col in range(1, 14):
            cell = ws.cell(row=row, column=col)
            apply_cell_style(cell, border, header_font, white_fill, wrap_center)

    apply_cell_style(ws["A1"], border, title_font, green_fill, center)
    apply_cell_style(ws["A2"], border, header_font, white_fill, center)
    apply_cell_style(ws["L2"], border, Font(name="Calibri", size=16, bold=True), white_fill, center)
    apply_cell_style(ws["D7"], border, product_font, product_fill, center)

    start_row = 8
    total_cft = Decimal("0")

    for idx, row in preview_df.iterrows():
        excel_row = start_row + idx
        component = row["Component"]
        cft = Decimal(str(row["CFT"] or 0))

        ws.cell(excel_row, 1, idx + 1)

        ws.merge_cells(
            start_row=excel_row,
            start_column=2,
            end_row=excel_row,
            end_column=3
        )
        ws.cell(excel_row, 2, component)

        ws.merge_cells(
            start_row=excel_row,
            start_column=5,
            end_row=excel_row,
            end_column=7
        )
        ws.cell(excel_row, 5, wood_shade_for_component(component))

        ws.cell(excel_row, 8, row["Length"])
        ws.cell(excel_row, 9, row["Width"])
        ws.cell(excel_row, 10, row["Thickness"])
        ws.cell(excel_row, 11, row["Qty"])
        ws.cell(excel_row, 12, float(cft) if cft != 0 else 0)
        ws.cell(excel_row, 13, row.get("LH & RH Details", ""))

        if not is_flush_component(component):
            total_cft += cft

        for col in range(1, 14):
            cell = ws.cell(excel_row, col)
            font = item_font if col in [2, 5] else normal_font
            apply_cell_style(cell, border, font, white_fill, center)

    total_row = start_row + len(preview_df)

    ws.merge_cells(
        start_row=total_row,
        start_column=1,
        end_row=total_row,
        end_column=11
    )
    ws.cell(total_row, 12, float(round(total_cft, 2)))
    ws.cell(total_row, 13, "")

    for col in range(1, 14):
        apply_cell_style(ws.cell(total_row, col), border, header_font, white_fill, center)

    widths = {
        "A": 8,
        "B": 20,
        "C": 20,
        "D": 16,
        "E": 20,
        "F": 20,
        "G": 20,
        "H": 12,
        "I": 12,
        "J": 12,
        "K": 10,
        "L": 12,
        "M": 38,
    }

    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    ws.row_dimensions[1].height = 30

    for row in range(2, total_row + 1):
        ws.row_dimensions[row].height = 28

    ws.freeze_panes = "A8"

    wb.save(output)
    output.seek(0)

    return output


def show_component_calculator(conn, cur):
    st.title("Component Calculator")

    projects = get_distinct_values(cur, "projects", "project_name")

    if not projects:
        st.info("No projects found.")
        return

    col1, col2 = st.columns(2)

    with col1:
        project_name = st.selectbox("Project", projects)

    unit_types = get_distinct_values(
        cur,
        "unit_types",
        "unit_type",
        "WHERE project_name = %s",
        (project_name,)
    )

    if not unit_types:
        st.info("No unit types found for this project.")
        return

    with col2:
        unit_type = st.selectbox("Unit Type", unit_types)

    houses = get_distinct_values(
        cur,
        "houses",
        "house_number",
        "WHERE project_name = %s AND unit_type = %s",
        (project_name, unit_type)
    )

    products_df = fetch_df(
        cur,
        """
        SELECT product_cat, product_code
        FROM products
        ORDER BY product_cat, product_code
        """
    )

    if products_df.empty:
        st.info("No products found.")
        return

    product_options = {
        f"{row['product_cat']} - {row['product_code']}": {
            "product_cat": row["product_cat"],
            "product_code": row["product_code"],
        }
        for _, row in products_df.iterrows()
    }

    col1, col2 = st.columns(2)

    with col1:
        selected_product = st.selectbox(
            "Product",
            list(product_options.keys())
        )

    with col2:
        selected_houses = st.multiselect(
            "House Number",
            houses
        )

    product_cat = product_options[selected_product]["product_cat"]
    product_code = product_options[selected_product]["product_code"]
    state_key = f"{project_name}_{unit_type}_{product_cat}_{product_code}"

    rules = load_product_rules(cur, product_cat, product_code)

    if not rules:
        st.warning("No product definition found for this product.")
        return

    uses_lh_rh = product_uses_lh_rh(product_cat, product_code, rules)

    house_quantities = {}
    total_lh_quantity = 0
    total_rh_quantity = 0
    total_product_quantity = 0

    if selected_houses:
        st.markdown("---")

        if uses_lh_rh:
            st.subheader("House Wise LH / RH Quantity")

            qty_df = pd.DataFrame({
                "House Number": selected_houses,
                "LH Quantity": [0 for _ in selected_houses],
                "RH Quantity": [0 for _ in selected_houses],
            })

            edited_qty_df = st.data_editor(
                qty_df,
                use_container_width=True,
                hide_index=True,
                disabled=["House Number"],
                key=f"lh_rh_qty_{state_key}",
            )

            for _, row in edited_qty_df.iterrows():
                house_number = row["House Number"]
                lh_qty = clean_int(row["LH Quantity"])
                rh_qty = clean_int(row["RH Quantity"])

                house_quantities[house_number] = {
                    "lh_quantity": lh_qty,
                    "rh_quantity": rh_qty,
                    "quantity": lh_qty + rh_qty,
                }

            total_lh_quantity = sum(item["lh_quantity"] for item in house_quantities.values())
            total_rh_quantity = sum(item["rh_quantity"] for item in house_quantities.values())
            total_product_quantity = total_lh_quantity + total_rh_quantity

            st.info(
                f"Total LH Quantity: {total_lh_quantity} | "
                f"Total RH Quantity: {total_rh_quantity} | "
                f"Total Quantity: {total_product_quantity}"
            )

        else:
            st.subheader("House Wise Quantity")

            qty_df = pd.DataFrame({
                "House Number": selected_houses,
                "Quantity": [1 for _ in selected_houses],
            })

            edited_qty_df = st.data_editor(
                qty_df,
                use_container_width=True,
                hide_index=True,
                disabled=["House Number"],
                key=f"product_qty_{state_key}",
            )

            for _, row in edited_qty_df.iterrows():
                house_number = row["House Number"]
                qty = clean_int(row["Quantity"])

                house_quantities[house_number] = {
                    "lh_quantity": 0,
                    "rh_quantity": 0,
                    "quantity": qty,
                }

            total_product_quantity = sum(item["quantity"] for item in house_quantities.values())

            st.info(f"Total Quantity: {total_product_quantity}")

    st.markdown("---")
    st.subheader("Required Inputs")

    formula_inputs = required_formula_inputs(rules)
    manual_inputs = required_manual_inputs(rules)

    user_inputs = {}

    if formula_inputs:
        st.markdown("#### Formula Inputs")
        formula_cols = st.columns(4)

        for idx, key in enumerate(formula_inputs):
            with formula_cols[idx % 4]:
                user_inputs[key] = st.number_input(
                    label_from_key(key),
                    value=0.0,
                    step=1.0,
                    format="%.2f",
                    key=f"formula_input_{state_key}_{key}"
                )

    if manual_inputs:
        st.markdown("#### Manual Inputs")
        manual_cols = st.columns(4)

        for idx, item in enumerate(manual_inputs):
            with manual_cols[idx % 4]:
                user_inputs[item["key"]] = st.number_input(
                    item["label"],
                    value=0.0,
                    step=1.0,
                    format="%.2f",
                    key=f"manual_input_{state_key}_{item['key']}"
                )

    if not formula_inputs and not manual_inputs:
        st.info("No user input required. Product uses only fixed values.")

    prepared_by = st.selectbox(
        "Prepared by",
        ["Anup", "Mani"],
        key=f"prepared_by_{state_key}"
    )

    st.markdown("---")

    if st.button("Generate Components", type="primary"):
        if not selected_houses:
            st.warning("Select at least one house.")
            return

        raw_preview_rows = []
        tracking_rows = []
        generated_insert_rows = []

        try:
            for house_number in selected_houses:
                house_qty = house_quantities.get(house_number, {
                    "lh_quantity": 0,
                    "rh_quantity": 0,
                    "quantity": 1,
                })

                product_qty = int(house_qty["quantity"])

                if product_qty <= 0:
                    st.warning(f"Enter quantity for house {house_number}.")
                    continue

                house_inputs = dict(user_inputs)
                house_inputs["lh_quantity"] = house_qty["lh_quantity"]
                house_inputs["rh_quantity"] = house_qty["rh_quantity"]
                house_inputs["quantity"] = product_qty

                calculated_rows, variables = calculate_rules(rules, house_inputs)

                summary_rows = build_component_summary(
                    rules,
                    calculated_rows,
                    product_qty
                )

                for row in summary_rows:
                    raw_preview_rows.append({
                        "House Number": house_number,
                        "Product": product_code,
                        "Component": row["Component"],
                        "Length": row["Length"],
                        "Width": row["Width"],
                        "Thickness": row["Thickness"],
                        "Qty": row["Qty"],
                        "CFT": row["CFT"],
                        "LH Quantity": house_qty["lh_quantity"],
                        "RH Quantity": house_qty["rh_quantity"],
                    })

                    generated_insert_rows.append((
                        project_name,
                        unit_type,
                        house_number,
                        product_cat,
                        product_code,
                        row["Component"],
                        "combined",
                        str(row["Length"]),
                        row["Width"],
                        row["Thickness"],
                        "",
                        row["Qty"],
                    ))

                    tracking_rows.append((
                        project_name,
                        house_number,
                        row["Component"],
                        row["Qty"],
                        0,
                        row["Qty"],
                        "Pending",
                    ))

            display_df = aggregate_preview_rows(raw_preview_rows)

            st.session_state["component_raw_preview_rows"] = raw_preview_rows
            st.session_state["component_preview_rows"] = display_df.to_dict("records")
            st.session_state["component_generated_rows"] = generated_insert_rows
            st.session_state["component_tracking_rows"] = tracking_rows
            st.session_state["component_total_lh_quantity"] = total_lh_quantity
            st.session_state["component_total_rh_quantity"] = total_rh_quantity
            st.session_state["component_total_quantity"] = total_product_quantity
            st.session_state["component_prepared_by"] = prepared_by
            st.session_state["component_opening_length"] = user_inputs.get("opening_length", "")
            st.session_state["component_opening_width"] = user_inputs.get("opening_width", "")

        except FormulaError as e:
            st.error(f"Calculation error: {e}")
            return

        except Exception as e:
            st.error(f"Error: {e}")
            return

    if "component_preview_rows" in st.session_state:
        st.subheader("Generated Components")

        display_df = pd.DataFrame(st.session_state["component_preview_rows"])

        total_lh_quantity = st.session_state.get("component_total_lh_quantity", 0)
        total_rh_quantity = st.session_state.get("component_total_rh_quantity", 0)
        total_product_quantity = st.session_state.get("component_total_quantity", 0)
        prepared_by = st.session_state.get("component_prepared_by", prepared_by)
        opening_length = st.session_state.get("component_opening_length", "")
        opening_width = st.session_state.get("component_opening_width", "")

        total_cft = sum(
            clean_number(value) or 0
            for value in display_df["CFT"].tolist()
        )

        st.info(
            f"Total LH Quantity: {total_lh_quantity} | "
            f"Total RH Quantity: {total_rh_quantity} | "
            f"Total Quantity: {total_product_quantity} | "
            f"Total CFT: {round(total_cft, 2)}"
        )

        total_row = {
            "Product": "",
            "Component": "CFT Total",
            "Length": "",
            "Width": "",
            "Thickness": "",
            "Qty": "",
            "CFT": Decimal(str(round(total_cft, 2))),
            "LH & RH Details": "",
        }

        display_with_total_df = pd.concat(
            [display_df, pd.DataFrame([total_row])],
            ignore_index=True
        )

        st.dataframe(
            display_with_total_df,
            use_container_width=True,
            hide_index=True
        )

        excel_file = build_wood_issue_excel(
            project_name=project_name,
            unit_type=unit_type,
            product_code=product_code,
            prepared_by=prepared_by,
            total_qty=total_product_quantity,
            total_lh_qty=total_lh_quantity,
            total_rh_qty=total_rh_quantity,
            opening_length=opening_length,
            opening_width=opening_width,
            preview_df=display_df,
        )

        st.download_button(
            label="Download Excel",
            data=excel_file,
            file_name=f"{project_name}_{unit_type}_{product_code}_components.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key=f"download_excel_{state_key}"
        )

        st.success("All component formulas calculated successfully")

        if st.button("Confirm Components and Send to Tracking", type="primary"):
            generated_rows = st.session_state.get("component_generated_rows", [])
            tracking_rows = st.session_state.get("component_tracking_rows", [])

            if not generated_rows:
                st.warning("No generated components to save.")
                return

            try:
                execute_values(
                    cur,
                    """
                    INSERT INTO generated_components
                    (
                        project_name,
                        unit_type,
                        house_number,
                        product_cat,
                        product_code,
                        component,
                        attribute,
                        calculated_value,
                        width,
                        thickness,
                        orientation,
                        qty
                    )
                    VALUES %s
                    """,
                    generated_rows
                )

                execute_values(
                    cur,
                    """
                    INSERT INTO tracking
                    (
                        project_name,
                        house_number,
                        component,
                        required_qty,
                        completed_qty,
                        pending_qty,
                        status
                    )
                    VALUES %s
                    """,
                    tracking_rows
                )

                conn.commit()
                st.success("Components sent to tracking successfully.")

            except Exception as e:
                conn.rollback()
                st.error(f"Failed to save generated components: {e}")
