from datetime import date
from io import BytesIO

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


def clean_number(value):
    if value in [None, ""]:
        return None

    try:
        if pd.isna(value):
            return None
    except Exception:
        pass

    try:
        return float(value)
    except Exception:
        return None


def clean_int(value):
    number = clean_number(value)

    if number is None:
        return 0

    return int(float(number))


def build_components_excel(
    project_name,
    unit_type,
    product_code,
    generated_lh,
    generated_rh,
    df_preview,
    project_code="",
    prepared_date=None,
    order_qty=None,
    opening_length="",
    opening_width="",
    prepared_by="",
):
    output = BytesIO()

    if prepared_date in [None, ""]:
        prepared_date = date.today().strftime("%d-%m-%Y")

    if prepared_by in [None, ""]:
        prepared_by = "Anup S Kalkundi"

    wb = Workbook()
    ws = wb.active
    ws.title = "Wood Issue Report"

    white_fill = PatternFill("solid", fgColor="FFFFFF")
    title_green = PatternFill("solid", fgColor="C6EFCE")
    grey_fill = PatternFill("solid", fgColor="D9E1DD")

    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    title_font = Font(bold=True, size=14, color="000000")
    header_font = Font(bold=True, size=11, color="000000")
    normal_font = Font(size=11, color="000000")
    item_font = Font(size=14, color="000000", name="Times New Roman")
    big_font = Font(bold=True, size=20, color="000000", name="Times New Roman")

    center = Alignment(horizontal="center", vertical="center")
    wrap_center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    generated_lh = clean_int(generated_lh)
    generated_rh = clean_int(generated_rh)

    total_qty = generated_lh + generated_rh

    if order_qty is None:
        order_qty = total_qty

    for row in ws.iter_rows(min_row=1, max_row=40, min_col=1, max_col=14):
        for cell in row:
            cell.fill = white_fill
            cell.border = border
            cell.alignment = center
            cell.font = normal_font

    ws.merge_cells("A1:N1")
    ws["A1"] = "WOOD ISSUE REPORT"
    ws["A1"].fill = title_green
    ws["A1"].font = title_font
    ws["A1"].alignment = center

    ws.merge_cells("A2:A6")
    ws["A2"] = "1"
    ws["A2"].font = header_font

    ws.merge_cells("B2:D2")
    ws["B2"] = "PROJECT NAME"
    ws.merge_cells("E2:G2")
    ws.merge_cells("H2:K2")
    ws["H2"] = f"{project_name} {unit_type} - {product_code} Batch".strip()

    ws.merge_cells("B3:D3")
    ws["B3"] = "PROJECT CODE"
    ws.merge_cells("E3:G3")
    ws.merge_cells("H3:K3")
    ws["H3"] = project_code

    ws.merge_cells("B4:D4")
    ws["B4"] = "PREPARED DATE"
    ws.merge_cells("E4:G4")
    ws.merge_cells("H4:K4")
    ws["H4"] = prepared_date

    ws.merge_cells("B5:D5")
    ws["B5"] = "ORDER qty"
    ws.merge_cells("E5:G5")
    ws["H5"] = order_qty

    lh_rh_text = []
    if generated_lh:
        lh_rh_text.append(f"{generated_lh} LH")
    if generated_rh:
        lh_rh_text.append(f"{generated_rh} RH")

    ws.merge_cells("I5:K5")
    ws["I5"] = " / ".join(lh_rh_text)

    ws.merge_cells("B6:D6")
    ws.merge_cells("E6:G6")
    ws.merge_cells("H6:I6")
    ws["H6"] = "OPENING SIZE"
    ws["J6"] = opening_length
    ws["K6"] = opening_width

    ws.merge_cells("L2:N4")
    ws["L2"] = total_qty
    ws["L2"].font = Font(bold=True, size=16)

    ws.merge_cells("L5:N5")
    ws["L5"] = "Prepared By"
    ws["L5"].font = header_font

    ws.merge_cells("L6:N6")
    ws["L6"] = prepared_by
    ws["L6"].font = header_font

    ws.merge_cells("B7:C7")
    ws.merge_cells("E7:G7")

    headers = {
        "A7": "SL.\nNO.",
        "B7": "Item Code",
        "E7": "WOOD SHADE",
        "H7": "LENGTH",
        "I7": "WIDTH",
        "J7": "THICK",
        "K7": "QTY",
        "L7": "CFT",
        "M7": "LH & RH DETAILS",
        "N7": "REMARKS",
    }

    for cell_ref, value in headers.items():
        ws[cell_ref] = value
        ws[cell_ref].font = header_font
        ws[cell_ref].alignment = wrap_center

    start_row = 8
    serial_no = 1
    total_cft = 0

    data_rows = df_preview[
        df_preview["Component"].astype(str).str.strip().str.lower() != "cft total"
    ]

    for _, row in data_rows.iterrows():
        excel_row = start_row + serial_no - 1

        component = row.get("Component", "")
        length = row.get("Length", "")
        width = row.get("Width", "")
        thickness = row.get("Thickness", "")
        qty = row.get("Total Quantity", "")
        cft = row.get("CFT", "")
        lh_rh_details = row.get("LH & RH Details", "")

        component_lower = str(component).strip().lower()

        if component_lower == "flush shutter":
            wood_shade = f"{thickness} mm BB" if thickness else "BB"
        else:
            wood_shade = "SOLIDWOOD - TEAK"

        ws.cell(excel_row, 1, serial_no)

        ws.merge_cells(start_row=excel_row, start_column=2, end_row=excel_row, end_column=3)
        ws.cell(excel_row, 2, component)

        ws.merge_cells(start_row=excel_row, start_column=5, end_row=excel_row, end_column=7)
        ws.cell(excel_row, 5, wood_shade)

        ws.cell(excel_row, 8, length)
        ws.cell(excel_row, 9, width)
        ws.cell(excel_row, 10, thickness)
        ws.cell(excel_row, 11, qty)
        ws.cell(excel_row, 12, cft)
        ws.cell(excel_row, 13, lh_rh_details)
        ws.cell(excel_row, 14, "")

        for col in range(1, 15):
            cell = ws.cell(excel_row, col)
            cell.fill = white_fill
            cell.border = border
            cell.alignment = wrap_center if col == 13 else center
            cell.font = item_font if col in [2, 5] else normal_font

        if clean_number(cft) is not None:
            total_cft += clean_number(cft)

        serial_no += 1

    last_data_row = start_row + len(data_rows) - 1

    if last_data_row >= start_row:
        ws.merge_cells(start_row=start_row, start_column=4, end_row=last_data_row, end_column=4)
        ws.cell(start_row, 4).value = product_code
        ws.cell(start_row, 4).fill = grey_fill
        ws.cell(start_row, 4).font = big_font
        ws.cell(start_row, 4).alignment = center

    total_row = start_row + len(data_rows)

    ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=11)
    ws.cell(total_row, 12, round(total_cft, 2))

    for col in range(1, 15):
        cell = ws.cell(total_row, col)
        cell.fill = white_fill
        cell.border = border
        cell.alignment = center
        cell.font = header_font if col == 12 else normal_font

    for cell_ref in ["H2", "H3", "H4", "H5", "I5", "H6", "J6", "K6", "L6"]:
        ws[cell_ref].font = header_font

    widths = {
        "A": 8,
        "B": 18,
        "C": 18,
        "D": 16,
        "E": 16,
        "F": 16,
        "G": 16,
        "H": 12,
        "I": 10,
        "J": 10,
        "K": 9,
        "L": 10,
        "M": 32,
        "N": 16,
    }

    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    ws.row_dimensions[1].height = 28

    for row_no in range(2, 7):
        ws.row_dimensions[row_no].height = 24

    ws.row_dimensions[7].height = 34

    for row_no in range(start_row, total_row + 1):
        ws.row_dimensions[row_no].height = 28

    ws.freeze_panes = "A7"

    wb.save(output)
    output.seek(0)

    return output
